"""
BET adatletolto - 3. lepes
Letolti minden ceg tobbeves penzugyi osszefoglalojat es fontosabb mutatoit
(a BET "Excel letoltes" linkje mogotti adatot), es cegenkent egy JSON
fajlba menti.

Fontos: a cegek nem egyforma sémát hasznalnak (pl. a bankoknal mas sorok
szerepelnek, mint egy tipikus cegnel), ezert altalanos "cimke -> ertek"
szerkezetben taroljuk az adatokat, nem fix oszlopokban.
"""

import io
import json
import os
import re
import time
import urllib.parse
import urllib.request
import warnings

import openpyxl

from bet_api import BetKapcsolat

warnings.filterwarnings("ignore")  # openpyxl stilus-figyelmeztetesek elnemitasa

KIMENET_MAPPA = "data/penzugy"


def bet_utvonalnev(kod):
    """
    A BET oldala a reszvenykodban levo '/' jelet NEM szabvanyos URL-kodolassal
    (%2F), hanem sajat, egyedi '0xc2F' helyettesitovel kezeli az utvonalaiban
    (pl. 'FORRAS/OE' -> 'FORRAS0xc2FOE'). Eloszor ezt cliseljuk le, utana
    MINDEN mas specialis karaktert (pl. szokoz) szabvanyos modon kodolunk.
    Ezt a fuggvenyt hasznaljuk mindenhol, ahol egy reszvenykodot utvonalba
    illesztunk - ezutan mar nem szabad ujra kodolni.
    """
    ideiglenes = kod.replace("/", "\x00")  # ideiglenes jelolo, hogy a quote ne bantsa
    kodolt = urllib.parse.quote(ideiglenes, safe="")
    return kodolt.replace("%00", "0xc2F")


def issuer_azonositok_frissitese(bet, reszvenyek):
    """
    Minden reszvenyhez kideriti a belso 'issuer id'-t a cegadatlap oldalarol.
    Ez kell ahhoz, hogy le tudjuk tolteni az Excel exportokat.
    """
    eredmeny = {}
    for i, r in enumerate(reszvenyek, 1):
        kod = r["seccode"]
        try:
            html = bet.oldal(f"/oldalak/ceg_adatlap/$security/{bet_utvonalnev(kod)}")
            issuer = re.search(r"issuer=(\d+)", html)
            eredmeny[kod] = {
                "securityid": r["securityid"],
                "issuerid": issuer.group(1) if issuer else None,
            }
        except Exception as hiba:
            eredmeny[kod] = {"securityid": r["securityid"], "issuerid": None, "hiba": str(hiba)}
        print(f"  [{i:2d}/{len(reszvenyek)}] {kod:14s} issuer={eredmeny[kod]['issuerid']}")
        time.sleep(0.3)
    return eredmeny


def egyseg_es_penznem(cimke_szoveg):
    """
    A BET Excel-exportban minden ev oszlop alatt ott all a mertekegyseg ES a
    penznem egyutt, pl. '(ezer HUF)' vagy '(ezer EUR)' - kulfoldi bevezetesu
    cegek (pl. VIG) nem HUF-ban, hanem sajat devizajukban jelentenek!
    Visszaadja: (szorzo, penznem) - pl. (1000, 'EUR').
    """
    if not cimke_szoveg:
        return 1, None
    szoveg = str(cimke_szoveg).lower()
    m = re.search(r'\(\s*(ezer|millió|milliárd)?\s*([a-z]{3})\s*\)', szoveg)
    if not m:
        return 1, None
    nagysagrend, penznem = m.groups()
    szorzo = {"milliárd": 1_000_000_000, "millió": 1_000_000, "ezer": 1_000}.get(nagysagrend, 1)
    return szorzo, penznem.upper()


def excel_letoltes_es_feldolgozas(bet, url):
    """Letolt egy Excel fajlt a BET-rol, es 'cimke -> {ev: ertek}' szotarra alakitja.
    Minden erteket tenyleges (nem ezerre/millora skalazott) HUF-ra normalizal,
    az Excelben evenkent kulon feltuntetett mertekegyseg alapjan."""
    nyers = None
    for proba in range(3):
        try:
            keres = urllib.request.Request("https://www.bet.hu" + url)
            keres.add_header("User-Agent",
                              "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
            nyers = bet.opener.open(keres, timeout=45).read()
            break
        except Exception:
            time.sleep(1.5 * (proba + 1))
    if nyers is None:
        raise RuntimeError("Nem sikerult letolteni az Excel fajlt")

    wb = openpyxl.load_workbook(io.BytesIO(nyers))
    ws = wb.active

    # Az evszamokat tartalmazo sor megkeresese (pl. "2024 (1)", "2025 (2)")
    ev_sor_index = None
    evek = []
    for sor_i in range(1, min(ws.max_row, 10) + 1):
        ertekek = [ws.cell(row=sor_i, column=c).value for c in range(2, ws.max_column + 1)]
        talalt_evek = []
        for e in ertekek:
            if e is None:
                continue
            m = re.match(r"^(\d{4})", str(e))
            if m:
                talalt_evek.append(m.group(1))
        if len(talalt_evek) >= 2:
            ev_sor_index = sor_i
            evek = talalt_evek
            break

    if ev_sor_index is None:
        return {"cim": None, "evek": [], "sorok": {}}

    cim = ws.cell(row=1, column=1).value

    # A mertekegyseg+penznem sor kozvetlenul az ev-sor alatt van, pl. "(ezer HUF)".
    # Oszloponkent kulon allapitjuk meg - igy akkor is jo, ha a ceg kozben
    # valtott mertekegyseget VAGY devizanemet.
    adat_kezdo_sor = ev_sor_index + 1
    szorzok = [1] * (len(evek) + 1)  # +1, mert 1-tol indexelunk oszlopban
    penznemek = {}  # ev -> penznem (pl. "HUF", "EUR")
    egyseg_sor_letezik = False
    for oszlop in range(2, 2 + len(evek)):
        cellaertek = ws.cell(row=ev_sor_index + 1, column=oszlop).value
        szorzo, penznem = egyseg_es_penznem(cellaertek)
        if penznem:
            egyseg_sor_letezik = True
            szorzok[oszlop - 1] = szorzo
            penznemek[evek[oszlop - 2]] = penznem
    if egyseg_sor_letezik:
        adat_kezdo_sor = ev_sor_index + 2

    sorok = {}
    for sor_i in range(adat_kezdo_sor, ws.max_row + 1):
        cimke = ws.cell(row=sor_i, column=1).value
        if not cimke or str(cimke).startswith("(") or "közölt információk" in str(cimke):
            continue
        cimke_tiszta = str(cimke).strip()
        # A "reszvenyre jutó" soroknak (EPS, osztalek/reszveny) MINDIG sima HUF-ban
        # adjak meg az erteket, fuggetlenul az oszlop mertekegyseg-cimkejetol -
        # ezekre sosem szabad a millio/ezer szorzot alkalmazni.
        per_reszveny_sor = "részvényre jutó" in cimke_tiszta
        ertekek_evenkent = {}
        oszlop = 2
        for ev in evek:
            ertek = ws.cell(row=sor_i, column=oszlop).value
            if ertek is not None and ertek != "":
                if isinstance(ertek, (int, float)) and not per_reszveny_sor:
                    ertek = ertek * szorzok[oszlop - 1]
                ertekek_evenkent[ev] = ertek
            oszlop += 1
        if ertekek_evenkent:
            sorok[cimke_tiszta] = ertekek_evenkent

    return {"cim": cim, "evek": evek, "sorok": sorok, "penznemek": penznemek}


def main():
    with open("data/reszvenyek.json", encoding="utf-8") as f:
        reszvenyek = json.load(f)["reszvenyek"]

    os.makedirs(KIMENET_MAPPA, exist_ok=True)
    bet = BetKapcsolat()

    print("Issuer azonositok lekerdezese...")
    issuer_map = issuer_azonositok_frissitese(bet, reszvenyek)
    with open("data/issuer_map.json", "w", encoding="utf-8") as f:
        json.dump(issuer_map, f, ensure_ascii=False, indent=2)

    print("\nPenzugyi adatok letoltese cegenkent...")
    hibas = []
    for i, r in enumerate(reszvenyek, 1):
        kod = r["seccode"]
        info = issuer_map.get(kod, {})
        issuer = info.get("issuerid")
        security = info.get("securityid")
        fajlnev = kod.replace("/", "_").replace(" ", "_")

        if not issuer:
            print(f"[{i:2d}/{len(reszvenyek)}] {kod:14s} KIHAGYVA (nincs issuer id)")
            hibas.append(kod)
            continue

        kod_utvonalhoz = bet_utvonalnev(kod)
        eves_url = (f"/oldalak/ceg_adatlap/$security/{kod_utvonalhoz}/$rspid0x117770x12/"
                    f"$riOSSZEFOGLALO0x1EVES0x1ADATOK?issuer={issuer}")
        mutato_url = (f"/oldalak/ceg_adatlap/$security/{kod_utvonalhoz}/$rspid0x117770x12/"
                      f"$riFONTOSABB0x1PENZUGYI0x1MUTATOK?security={security}")

        try:
            eves_adat = excel_letoltes_es_feldolgozas(bet, eves_url)
            time.sleep(0.5)
            mutato_adat = excel_letoltes_es_feldolgozas(bet, mutato_url)
        except Exception as hiba:
            print(f"[{i:2d}/{len(reszvenyek)}] {kod:14s} HIBA: {hiba}")
            hibas.append(kod)
            time.sleep(1)
            continue

        kimenet = {
            "seccode": kod,
            "issuerid": issuer,
            "securityid": security,
            "eves_osszefoglalo": eves_adat,
            "penzugyi_mutatok": mutato_adat,
        }

        with open(f"{KIMENET_MAPPA}/{fajlnev}.json", "w", encoding="utf-8") as f:
            json.dump(kimenet, f, ensure_ascii=False, indent=2)

        ev_szam = len(eves_adat["evek"])
        sor_szam = len(eves_adat["sorok"])
        print(f"[{i:2d}/{len(reszvenyek)}] {kod:14s} {ev_szam} ev, {sor_szam} sor")
        time.sleep(0.6)

    print(f"\nKESZ. {len(reszvenyek) - len(hibas)}/{len(reszvenyek)} ceg sikeres.")
    if hibas:
        print("Kimaradt:", ", ".join(hibas))


if __name__ == "__main__":
    main()
